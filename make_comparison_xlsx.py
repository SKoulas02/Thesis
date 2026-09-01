"""Build GEMV_Hardware_Results.xlsx from the measured CSVs.

All measured on the physical Alveo U280, both designs validated bit-exact by the
same Python golden model from behavioural simulation through to silicon.

    dense_300MHz.csv                    dense_gemv  @ 300 MHz
    sparse_2to{4,8,16,32}_300MHz.csv    two2N       @ 300 MHz  <- HEADLINE
    sparse_2to{4,8,16,32}_225MHz.csv    two2N       @ 225 MHz  <- second point
    power_results.csv                   power/energy, all five @ 300 MHz
    power_results_325.csv               power/energy, all five @ 325 MHz

WHY TWO SPARSE CLOCKS. Sparse originally closed only 225 MHz against dense's 300.
Two changes fixed that -- per-core replication of the activation-window register
(`equivalent_register_removal` + `max_fanout` on `A_internal`) and an SLR
floorplan (`gemv:SLR1`, all 17 movers `SLR0`) -- after which the SAME RTL closes
300 MHz. The 225 MHz set is kept because "floorplanning recovered a 33% clock
deficit" is itself a result, and because the pair together show the design's
throughput scales with clock as expected.

The 300 MHz set is the headline: with both designs at the same clock there is no
clock term in the comparison, so the speed-up IS the beat-count ratio.

WHY 325 MHz IS A THIRD POINT AND NOT A NEW HEADLINE. 325 is the highest clock
BOTH designs close (dense WNS +0.013, sparse +0.001), so it is the fairest
possible equal-clock comparison -- but the speed-up ratios are clock-independent
by construction, so it CONFIRMS 300 rather than replacing it. Its throughput
figures come from the 60 s power soaks, not from marginal fits: the 4-point
sweeps at 325 were run on a loaded machine and one of them (2:4) came back
PHYSICALLY IMPOSSIBLE -- 1.004 beats/cycle, faster than the clock itself. See
the '325 MHz' sheet, section 5. NEVER compare a 325 soak number against a 300
fit number: different estimators, ~6% apart, and the difference is not physical.

Run:  python make_comparison_xlsx.py
"""

import csv
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "GEMV_Hardware_Results.xlsx")

DENSE = ("dense", "dense_300MHz.csv", 300.0)
SPARSE300 = [("2:4", "sparse_2to4_300MHz.csv"), ("2:8", "sparse_2to8_300MHz.csv"),
             ("2:16", "sparse_2to16_300MHz.csv"), ("2:32", "sparse_2to32_300MHz.csv")]
SPARSE225 = [("2:4", "sparse_2to4_225MHz.csv"), ("2:8", "sparse_2to8_225MHz.csv"),
             ("2:16", "sparse_2to16_225MHz.csv"), ("2:32", "sparse_2to32_225MHz.csv")]

# Independent repeat runs at 300 MHz, from the terminal output. The CSVs were
# overwritten by the higher-rep reruns, so these are recorded here as replicate
# evidence: two separate measurements agreeing is a better uncertainty estimate
# than one fit's R^2.
REPLICATES = {"2:8": (145.831, 0.99876), "2:16": (294.688, 0.99877)}

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(color="FFFFFF", bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
HI_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
TITLE = Font(bold=True, size=14, color="1F3864")
SEC = Font(bold=True, size=11, color="1F3864")
NOTE = Font(italic=True, size=9, color="555555")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="AAAAAA")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(path):
    with io.open(os.path.join(HERE, path), encoding="utf-8") as f:
        return list(csv.DictReader(f))


def fnum(row, key):
    v = row.get(key, "")
    if v in ("", None):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def digest(path):
    rows = load(path)
    fit = [r for r in rows if r["laps"] == "FIT"][0]
    sizes = [r for r in rows if r["laps"].isdigit()]
    big = sizes[-1]
    return dict(
        rows=rows,
        overhead_us=fnum(fit, "best_us"), slope_us=fnum(fit, "mean_us"),
        r2=fnum(fit, "worst_us"), mrow=fnum(fit, "Mrow_s"),
        gbs=fnum(fit, "GB_s"), bpc=fnum(fit, "beats_per_cycle"),
        beats_per_row=float(big["weight_beats"]) / float(big["rows"]),
        bytes_per_row=float(big["bytes_in"]) / float(big["rows"]),
        index_share=((fnum(big, "bytes_index") or 0.0) / float(big["bytes_in"])),
        n_sizes=int(float(fit["reps"])),
        beat_lo=int(sizes[0]["weight_beats"]), beat_hi=int(big["weight_beats"]),
    )


def header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill, cell.font = H_FILL, H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def put(ws, r, c, v, fmt=None, font=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.border = BOX
    return cell


def para(ws, r, text, font=None, width=8, height=28):
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    ws.row_dimensions[r].height = height
    return r + 1


D = digest(DENSE[1])
S3 = {lbl: digest(p) for lbl, p in SPARSE300}
S2 = {lbl: digest(p) for lbl, p in SPARSE225}

wb = Workbook()

# ===========================================================================
# 1 -- Comparison (300 MHz, equal clock)
# ===========================================================================
ws = wb.active
ws.title = "Comparison"
ws["A1"] = "GEMV on Alveo U280 - dense vs 2:M sparse, BOTH AT 300 MHz"
ws["A1"].font = TITLE
r = 2
r = para(ws, r, "Marginal (launch-overhead-free) figures from a least-squares fit over four "
                "problem sizes spanning an identical 262,144 - 2,097,152 weight-beat range. "
                "Both designs bit-exact on hardware, validated by the same golden model. "
                "With both at the same clock there is no clock term: the speed-up IS the "
                "beat-count ratio.", NOTE, height=30)

r += 1
ws.cell(row=r, column=1, value="1. HEADLINE - throughput at equal clock").font = SEC
r += 1
header(ws, r, ["Design", "Clock (MHz)", "Weight beats / row", "Throughput (Mrow/s)",
               "Speed-up vs dense", "Theory", "% of theory", "beats / cycle"],
       widths=[13, 11, 12, 15, 12, 9, 11, 11])
r += 1
for lbl in ["dense", "2:4", "2:8", "2:16", "2:32"]:
    f = D if lbl == "dense" else S3[lbl]
    name = "dense" if lbl == "dense" else "sparse " + lbl
    put(ws, r, 1, name, font=BOLD)
    put(ws, r, 2, 300, "0")
    put(ws, r, 3, f["beats_per_row"], "0.0##")
    put(ws, r, 4, f["mrow"], "0.00", font=BOLD, fill=(HI_FILL if lbl != "dense" else None))
    if lbl == "dense":
        put(ws, r, 5, 1.0, '0.00"x"')
        put(ws, r, 6, "")
        put(ws, r, 7, "")
    else:
        meas = f["mrow"] / D["mrow"]
        theory = D["beats_per_row"] / f["beats_per_row"]
        put(ws, r, 5, meas, '0.00"x"', font=BOLD, fill=HI_FILL)
        put(ws, r, 6, theory, '0"x"')
        put(ws, r, 7, meas / theory, "0.0%", fill=OK_FILL)
    put(ws, r, 8, f["bpc"], "0.000")
    r += 1

r += 1
ws.cell(row=r, column=1, value="2. COST PER WEIGHT BEAT - the architectural claim").font = SEC
r += 1
r = para(ws, r, "If the engine consumes one weight beat per cycle regardless of sparsity, "
                "the cost per beat must be constant AND equal to dense's. It is: 0.003379 "
                "(dense) against 0.003406-0.003459 us, within 2.4%. Sparsity changes only "
                "how many beats a row needs.", NOTE)
header(ws, r, ["Design", "us per weight beat", "vs dense", "Fixed overhead (us)",
               "R^2", "Sizes in fit", "Beat range"],
       widths=[13, 15, 11, 15, 11, 11, 22])
r += 1
for lbl in ["dense", "2:4", "2:8", "2:16", "2:32"]:
    f = D if lbl == "dense" else S3[lbl]
    put(ws, r, 1, "dense" if lbl == "dense" else "sparse " + lbl, font=BOLD)
    put(ws, r, 2, f["slope_us"], "0.000000", font=BOLD)
    put(ws, r, 3, f["slope_us"] / D["slope_us"], "0.0%")
    put(ws, r, 4, f["overhead_us"], "0.0")
    put(ws, r, 5, f["r2"], "0.00000", fill=(None if f["r2"] >= 0.999 else HI_FILL))
    put(ws, r, 6, f["n_sizes"], "0")
    put(ws, r, 7, "{:,} - {:,}".format(f["beat_lo"], f["beat_hi"]))
    r += 1

r += 1
ws.cell(row=r, column=1, value="3. DATA MOVED - the interface cost of sparsity").font = SEC
r += 1
r = para(ws, r, "Sparse carries 3 index pseudo-channels on top of 8 weight PCs: 11 x 256 "
                "bits per beat against dense's 8, i.e. 37.5% MORE bytes per beat. It needs "
                "far fewer beats, so traffic per row still collapses.", NOTE)
header(ws, r, ["Design", "Input PCs", "Bytes / beat", "Bytes / row", "vs dense",
               "Index share of input", "Bandwidth (GB/s)"],
       widths=[13, 10, 11, 12, 11, 14, 14])
r += 1
for lbl in ["dense", "2:4", "2:8", "2:16", "2:32"]:
    f = D if lbl == "dense" else S3[lbl]
    pcs = 8 if lbl == "dense" else 11
    put(ws, r, 1, "dense" if lbl == "dense" else "sparse " + lbl, font=BOLD)
    put(ws, r, 2, pcs, "0")
    put(ws, r, 3, pcs * 32, "0")
    put(ws, r, 4, f["bytes_per_row"], "0")
    put(ws, r, 5, f["bytes_per_row"] / D["bytes_per_row"], "0.0%", fill=HI_FILL)
    put(ws, r, 6, f["index_share"], "0.0%")
    put(ws, r, 7, f["gbs"], "0.00")
    r += 1
r += 1
r = para(ws, r, "Input-channel ceiling at 300 MHz: dense 8 x 32 B x 300 MHz = 76.8 GB/s; "
                "sparse 11 x 32 B x 300 MHz = 105.6 GB/s. Dense reaches 75.83 (98.7%), "
                "sparse ~103 (97.5-98.0%). Both essentially saturate their input channels.",
         NOTE)

r += 1
for line in [
    ("KEY FINDINGS", BOLD, 15),
    ("1. At equal clock the speed-up is the beat-count ratio, measured at 97.7-99.2% of "
     "theory: 1.98x / 3.95x / 7.86x / 15.63x against 2 / 4 / 8 / 16.", None, 28),
    ("2. Cost per weight beat is CONSTANT and matches dense within 2.4%. The engine takes "
     "one beat per cycle regardless of sparsity or design; sparsity only changes how many "
     "beats a row costs. This is the architecture's central claim, measured.", None, 40),
    ("3. beats/cycle stays near roofline at every ratio (0.964-0.979). There is NO "
     "output-path bottleneck at high sparsity - an earlier, noisier sweep suggested one and "
     "was wrong; see Methodology.", None, 40),
    ("4. Sparse wins at EVERY ratio including 2:4, where it halves the beats. The index "
     "overhead (37.5% more bytes per beat) is more than repaid.", None, 28),
    ("5. Both designs saturate their input channels (97.5-98.7% of the PC ceiling), so both "
     "are running at their architectural roofline, not against a tool or memory limit.",
     None, 28),
]:
    r = para(ws, r, line[0], line[1] or Font(size=10), height=line[2])

# ===========================================================================
# 2 -- Clock study
# ===========================================================================
ws = wb.create_sheet("Clock study")
ws["A1"] = "How sparse got from 225 MHz to 300 MHz - same RTL, two changes"
ws["A1"].font = TITLE
r = 2
r = para(ws, r, "Sparse originally missed 300 MHz badly. The fix was not architectural: two "
                "synthesis attributes plus a floorplan. Recorded because the intermediate "
                "numbers identify the mechanism.", NOTE)

r += 1
ws.cell(row=r, column=1, value="1. BUILD PROGRESSION (post-route, kernel clock)").font = SEC
r += 1
header(ws, r, ["Build", "Change from previous", "WNS (ns)", "TNS (ns)",
               "Failing endpoints", "Closes?"], widths=[22, 34, 11, 12, 14, 10])
r += 1
for name, change, wns, tns, fail, ok in [
    ("300 baseline", "-", -0.480, -737.4, 3263, "no"),
    ("300 + fanout fix", "equivalent_register_removal + max_fanout on A_internal",
     -0.299, -138.8, 1201, "no"),
    ("225", "lower target only", 0.048, 0.0, 0, "yes"),
    ("300 + fix + floorplan", "slr=gemv:SLR1, all 17 movers SLR0", 0.000, 0.0, 0, "YES"),
]:
    put(ws, r, 1, name, font=BOLD)
    put(ws, r, 2, change)
    put(ws, r, 3, wns, "0.000", fill=(OK_FILL if wns >= 0 else None))
    put(ws, r, 4, tns, "0.0")
    put(ws, r, 5, fail, "#,##0")
    put(ws, r, 6, ok, font=(BOLD if ok == "YES" else None),
        fill=(OK_FILL if ok != "no" else None))
    r += 1
r += 1
r = para(ws, r, "The fanout fix: all 8 cores latch the same activation window under the same "
                "enable, so Vivado merged the 8 per-core registers into ONE driving all 64 "
                "blocks - measured fanout 128 (= 64 blocks x 2 gather muxes), 3.246 ns on a "
                "single net, 93% of the critical path's route delay. Keeping the copies "
                "apart cut that to 16 loads each. Cost: ~3,584 FFs of 2.6 M.", NOTE, height=44)
r = para(ws, r, "The floorplan: gemv (66,186 LUTs) moved wholesale SLR0 -> SLR1, confirmed by "
                "the LUT deltas (-65,422 / +66,271). SLR1<->SLR0 crossings rose 2,554 -> "
                "6,853, i.e. +4,299 against a prediction of ~4,450 for the 17 AXIS streams. "
                "ADDING 4,299 crossings IMPROVED timing, because AXI4-Stream is "
                "handshake-based and latency-tolerant. SLR2 is platform and cannot be moved.",
         NOTE, height=44)

r += 1
ws.cell(row=r, column=1, value="2. THROUGHPUT AT BOTH CLOCKS").font = SEC
r += 1
r = para(ws, r, "The design should scale with clock. Measured ratio vs the ideal 1.333:",
         NOTE, height=15)
header(ws, r, ["Sparsity", "225 MHz (Mrow/s)", "300 MHz (Mrow/s)", "Ratio",
               "Ideal 300/225", "Deviation", "beats/cyc 225", "beats/cyc 300"],
       widths=[11, 15, 15, 10, 12, 11, 12, 12])
r += 1
for lbl in ["2:4", "2:8", "2:16", "2:32"]:
    a, b = S2[lbl], S3[lbl]
    put(ws, r, 1, "sparse " + lbl, font=BOLD)
    put(ws, r, 2, a["mrow"], "0.00")
    put(ws, r, 3, b["mrow"], "0.00", font=BOLD)
    put(ws, r, 4, b["mrow"] / a["mrow"], "0.000")
    put(ws, r, 5, 300.0 / 225.0, "0.000")
    put(ws, r, 6, (b["mrow"] / a["mrow"]) / (300.0 / 225.0) - 1.0, "+0.0%;-0.0%")
    put(ws, r, 7, a["bpc"], "0.000")
    put(ws, r, 8, b["bpc"], "0.000")
    r += 1
r += 1
r = para(ws, r, "All four within 2% of the ideal clock ratio, and beats/cycle is preserved - "
                "so the 33% extra clock converts directly into 33% more throughput. The "
                "engine was never the limit at 225 MHz; the system build was.", NOTE)

r += 1
ws.cell(row=r, column=1, value="3. REPLICATE RUNS (independent repeats at 300 MHz)").font = SEC
r += 1
r = para(ws, r, "Two of the sparsities were measured twice, independently. Agreement between "
                "repeats is a better uncertainty estimate than any single fit's R^2.", NOTE,
         height=15)
header(ws, r, ["Sparsity", "run 1 (Mrow/s)", "run 2 (Mrow/s)", "mean", "spread",
               "run 1 R^2", "run 2 R^2"], widths=[11, 14, 14, 12, 10, 11, 11])
r += 1
for lbl, (m1, r2_1) in REPLICATES.items():
    m2 = S3[lbl]["mrow"]
    put(ws, r, 1, "sparse " + lbl, font=BOLD)
    put(ws, r, 2, m1, "0.00")
    put(ws, r, 3, m2, "0.00")
    put(ws, r, 4, (m1 + m2) / 2.0, "0.00", font=BOLD)
    put(ws, r, 5, abs(m2 - m1) / ((m1 + m2) / 2.0), "0.0%", fill=OK_FILL)
    put(ws, r, 6, r2_1, "0.00000")
    put(ws, r, 7, S3[lbl]["r2"], "0.00000")
    r += 1
r += 1
r = para(ws, r, "The CSVs hold run 2 (higher rep count). Run 1 values are from the terminal "
                "record. 2:16 sits marginally below R^2 0.999 in both runs - the cause is the "
                "SMALLEST size in each sweep, where launch-overhead variance gives 25-32% "
                "spread against 4-5% at the largest. More repetitions do not fix that; the "
                "1.4% replicate agreement is the honest uncertainty.", NOTE, height=44)

r += 1
ws.cell(row=r, column=1, value="4. FMAX - how high each design actually goes").font = SEC
r += 1
r = para(ws, r, "Later builds pushed past 300. 325 closes on BOTH designs and is verified "
                "bit-exact on hardware at all four sparsities. 350 closes on dense only. The "
                "350 builds are the first in which a SYNTHESIS strategy actually applied "
                "(Flow_AlternateRoutability); at 300 and 325 no synthesis strategy applied at "
                "all, which is why those two clocks share one area result.", NOTE, height=44)
header(ws, r, ["Target clock", "Dense WNS (ns)", "Dense", "Sparse WNS (ns)", "Sparse",
               "Equal-clock comparison possible?"],
       widths=[12, 14, 10, 14, 10, 30])
r += 1
for tgt, dw, dok, sw, sok, note in [
        (225, None, "-", 0.048, "closes", "no - dense was never built here"),
        (300, 0.022, "closes", 0.000, "closes", "YES - the headline set"),
        (325, 0.013, "closes", 0.001, "closes", "YES - power/energy measured, see '325 MHz'"),
        (350, 0.025, "closes", -0.204, "FAILS", "no - dense only")]:
    put(ws, r, 1, tgt, "0")
    put(ws, r, 2, dw if dw is not None else "", "0.000",
        fill=(OK_FILL if dw is not None and dw >= 0 else None))
    put(ws, r, 3, dok, font=(BOLD if dok == "closes" else None))
    put(ws, r, 4, sw, "0.000", fill=(OK_FILL if sw >= 0 else None))
    put(ws, r, 5, sok, font=BOLD, fill=(None if sok == "closes" else HI_FILL))
    put(ws, r, 6, note)
    r += 1
r += 1
r = para(ws, r, "DENSE 350, SPARSE 325 - a residual gap of about 7%. Worth stating as the "
                "arc: sparse began 25% behind (it missed 300 by -0.480 ns while dense closed "
                "it). Register replication and the SLR floorplan removed most of that deficit; "
                "what survives is the gather's irreducible cost. The sparse 350 failure has an "
                "achieved period of 3.061 ns (~327 MHz), which corroborates the 325 that "
                "closed - but a failed run's WNS only ESTIMATES a ceiling. Fmax is established "
                "by a constraint that closes, never by one that does not. Measured "
                "counterexample on this very design: a 300 MHz target achieved 3.813 ns while "
                "a 250 MHz target achieved 4.177 ns.", NOTE, height=72)
r = para(ws, r, "This is a SEPARATE claim from the throughput result and does not weaken it: "
                "at equal clock sparse is 1.98x to 15.63x faster; dense sustains a 7% higher "
                "maximum clock. Both measured, neither hiding the other.", NOTE, height=30)

# ===========================================================================
# 3 -- Power and energy
# ===========================================================================
PWR = os.path.join(HERE, "power_results.csv")
if os.path.exists(PWR):
    praw = load("power_results.csv")

    # Average any repeated label -- a re-run appends rather than replaces, which
    # turns repeats into replicates. 2:4 was measured twice.
    order, grouped = [], {}
    for row in praw:
        lbl = row["label"]
        if lbl not in grouped:
            grouped[lbl] = []
            order.append(lbl)
        grouped[lbl].append(row)

    def avg(lbl, key):
        vals = [fnum(r, key) for r in grouped[lbl] if fnum(r, key) is not None]
        return sum(vals) / len(vals) if vals else None

    ws = wb.create_sheet("Power and energy")
    ws["A1"] = "Power and energy efficiency - all five configurations at 300 MHz"
    ws["A1"].font = TITLE
    r = 2
    r = para(ws, r, "60 s sustained load per configuration, buffers resident on the card so "
                    "the accelerator does compute and HBM traffic only -- no transfers. "
                    "Sampled at ~1 Hz via xbutil, clipped to the load window the host "
                    "reports, first 6 s discarded because the card's telemetry lags the "
                    "load. Idle baseline taken after the host exits with the SAME bitstream "
                    "still programmed, so leakage and clock trees are held constant.", NOTE,
             height=44)

    r += 1
    ws.cell(row=r, column=1, value="1. POWER - measured, with static and dynamic separated").font = SEC
    r += 1
    r = para(ws, r, "Board = the two 12 V input rails (FPGA + HBM + transceivers + regulator "
                    "losses + fans). VCCINT = the FPGA core rail, a SUBSET of board power "
                    "downstream of the regulators -- never add the two. Idle here means the "
                    "bitstream is programmed but nothing is running, so load - idle is the "
                    "power of doing the work.", NOTE, height=30)
    header(ws, r, ["Design", "Board load (W)", "Board idle (W)", "Board dynamic (W)",
                   "VCCINT load (W)", "VCCINT idle (W)", "VCCINT dynamic (W)",
                   "Load samples"],
           widths=[13, 13, 13, 14, 13, 13, 15, 11])
    r += 1
    for lbl in ["dense", "sparse 2:4", "sparse 2:8", "sparse 2:16", "sparse 2:32"]:
        if lbl not in grouped:
            continue
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, avg(lbl, "board_load_w"), "0.00")
        put(ws, r, 3, avg(lbl, "board_idle_w"), "0.00")
        put(ws, r, 4, avg(lbl, "board_delta_w"), "0.00", fill=SUB_FILL)
        put(ws, r, 5, avg(lbl, "vccint_load_w"), "0.00")
        put(ws, r, 6, avg(lbl, "vccint_idle_w"), "0.00")
        put(ws, r, 7, avg(lbl, "vccint_delta_w"), "0.00", fill=SUB_FILL)
        put(ws, r, 8, sum(int(x["n_load"]) for x in grouped[lbl]), "0")
        r += 1
    r += 1
    r = para(ws, r, "SPARSE POWER IS FLAT ACROSS SPARSITY: board load 35.33-35.52 W, dynamic "
                    "5.19-5.38 W, at every ratio from 2:4 to 2:32. That is the mechanism, not "
                    "a null result -- the engine consumes one weight beat per cycle whatever "
                    "the sparsity, so it draws the same power. Throughput rises 8x because a "
                    "row needs 8x fewer beats, while the watts stay put.", NOTE, height=44)
    r = para(ws, r, "SPARSE PAYS ~3.2 W OF STATIC OVERHEAD: idle 30.14 W against dense's "
                    "26.92. That is 66k LUTs against 44k plus interconnect for 17 HBM "
                    "channels instead of 14 -- paid whether the design is working or not, and "
                    "it is what dilutes the board-level energy advantage below the "
                    "throughput ratio.", NOTE, height=30)

    r += 1
    ws.cell(row=r, column=1, value="2. ENERGY EFFICIENCY - the result").font = SEC
    r += 1
    header(ws, r, ["Design", "Sustained (Mrow/s)", "Energy/row, board (nJ)", "vs dense",
                   "Energy/row, VCCINT (nJ)", "vs dense", "GMAC/s/W", "Throughput vs dense"],
           widths=[13, 15, 16, 11, 16, 11, 12, 14])
    r += 1
    d_mrow = avg("dense", "Mrow_s_sustained")
    d_nj = avg("dense", "energy_per_row_nJ")
    d_njv = avg("dense", "vccint_load_w") * 1000.0 / d_mrow
    for lbl in ["dense", "sparse 2:4", "sparse 2:8", "sparse 2:16", "sparse 2:32"]:
        if lbl not in grouped:
            continue
        m = avg(lbl, "Mrow_s_sustained")
        nj = avg(lbl, "energy_per_row_nJ")
        njv = avg(lbl, "vccint_load_w") * 1000.0 / m
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, m, "0.00")
        put(ws, r, 3, nj, "0.0", font=BOLD)
        put(ws, r, 4, d_nj / nj, '0.00"x"', fill=(HI_FILL if lbl != "dense" else None))
        put(ws, r, 5, njv, "0.0", font=BOLD)
        put(ws, r, 6, d_njv / njv, '0.00"x"', fill=(OK_FILL if lbl != "dense" else None))
        put(ws, r, 7, avg(lbl, "GMAC_per_W"), "0.00")
        put(ws, r, 8, m / d_mrow, '0.00"x"')
        r += 1
    r += 1
    r = para(ws, r, "READ THE TWO RAILS SEPARATELY -- they say different things and both "
                    "belong in the thesis. On VCCINT the energy-per-row gain (1.97 / 3.90 / "
                    "7.64 / 15.12x) tracks the throughput gain (1.98 / 3.96 / 7.86 / 15.63x) "
                    "almost exactly: on the FPGA core, sparsity converts to energy efficiency "
                    "one-for-one. On board power it is 1.75 / 3.50 / 6.91 / 13.63x -- still "
                    "large, but diluted by the 3.2 W static overhead sparse carries. The "
                    "board figure is what someone deploying this would actually pay.", NOTE,
             height=58)
    r = para(ws, r, "COUNTERINTUITIVE, AND OFFERED AS A HYPOTHESIS RATHER THAN A FINDING: "
                    "dense's VCCINT DYNAMIC power is 1.07 W against sparse's ~0.66 W -- dense "
                    "burns more core power while working, despite both doing 128 MACs per "
                    "cycle. A plausible mechanism is that sparse FREEZES its 512-bit "
                    "activation window for 32/M cycles, so those registers toggle up to 8x "
                    "less often, while dense's window advances every beat. Nothing measured "
                    "here isolates that, so it should not be presented as established.", NOTE,
             height=58)

    r += 1
    ws.cell(row=r, column=1, value="3. REPLICATE (2:4 measured twice, independently)").font = SEC
    r += 1
    header(ws, r, ["Quantity", "run 1", "run 2", "spread"], widths=[26, 12, 12, 10])
    r += 1
    g = grouped.get("sparse 2:4", [])
    if len(g) >= 2:
        for name, key, fmt in [("Sustained Mrow/s", "Mrow_s_sustained", "0.000"),
                               ("Board load (W)", "board_load_w", "0.000"),
                               ("Board dynamic (W)", "board_delta_w", "0.000"),
                               ("VCCINT dynamic (W)", "vccint_delta_w", "0.000"),
                               ("Energy per row (nJ)", "energy_per_row_nJ", "0.00")]:
            a1, a2 = fnum(g[0], key), fnum(g[1], key)
            put(ws, r, 1, name, font=BOLD)
            put(ws, r, 2, a1, fmt)
            put(ws, r, 3, a2, fmt)
            put(ws, r, 4, abs(a2 - a1) / ((a1 + a2) / 2.0), "0.0%", fill=OK_FILL)
            r += 1
    r += 1
    r = para(ws, r, "The tables above average the two 2:4 runs. Energy per row agrees to 0.4%; "
                    "the dynamic-power deltas, being small differences of large numbers, are "
                    "the noisiest quantity here and should be quoted with their standard "
                    "errors (in the raw CSV).", NOTE, height=30)

    r += 1
    for line in [("KEY FINDINGS", BOLD, 15),
                 ("1. Sparse power is INDEPENDENT of sparsity (35.33-35.52 W board at every "
                  "ratio) because the engine takes one weight beat per cycle regardless. "
                  "Energy per row therefore falls exactly as throughput rises.", None, 40),
                 ("2. Energy per row improves up to 13.6x (board) / 15.1x (VCCINT) at 2:32 "
                  "against dense at the same clock.", None, 28),
                 ("3. Sparse costs ~3.2 W of STATIC power for being a larger design with 3 "
                  "more HBM channels. That is the honest cost of the architecture and it is "
                  "why the board-level gain trails the throughput gain.", None, 40),
                 ("4. The efficiency advantage is larger than the throughput advantage on "
                  "VCCINT and smaller on board power -- quote which rail you mean.", None, 28)]:
        r = para(ws, r, line[0], line[1] or Font(size=10), height=line[2])

    # raw power rows, unaveraged
    ws2 = wb.create_sheet("power_raw")
    ws2.cell(row=1, column=1, value="power_results.csv - raw, one row per run, "
                                    "nothing averaged").font = TITLE
    pcols = ["label", "Mrow_s_sustained", "iterations", "board_load_w", "board_idle_w",
             "board_delta_w", "board_delta_se", "board_load_sd", "vccint_load_w",
             "vccint_idle_w", "vccint_delta_w", "vccint_delta_se", "n_load", "n_idle",
             "energy_per_row_nJ", "energy_per_MAC_pJ", "GMAC_per_W", "sample_errors"]
    header(ws2, 3, pcols, widths=[13] + [12] * (len(pcols) - 1))
    rr = 4
    for row in praw:
        for i, k in enumerate(pcols, start=1):
            if k == "label":
                put(ws2, rr, i, row.get(k, ""), font=BOLD)
            else:
                v = fnum(row, k)
                put(ws2, rr, i, v if v is not None else "",
                    "#,##0" if k == "iterations" else "0.000")
        rr += 1

# ===========================================================================
# 3b -- 325 MHz: the highest clock BOTH designs close
# ===========================================================================
PWR325 = os.path.join(HERE, "power_results_325.csv")
if os.path.exists(PWR325) and os.path.exists(PWR):
    XFMT = '0.00"x"'
    p325raw = load("power_results_325.csv")
    g325 = {}
    for row in p325raw:
        g325.setdefault(row["label"], []).append(row)

    def avg325(lbl, key):
        vals = [fnum(x, key) for x in g325.get(lbl, []) if fnum(x, key) is not None]
        return sum(vals) / len(vals) if vals else None

    ORDER = ["dense", "sparse 2:4", "sparse 2:8", "sparse 2:16", "sparse 2:32"]
    THEORY = {"sparse 2:4": 2.0, "sparse 2:8": 4.0, "sparse 2:16": 8.0, "sparse 2:32": 16.0}
    SPARSE_L = [l for l in ORDER[1:] if l in g325]

    ws = wb.create_sheet("325 MHz")
    ws["A1"] = "325 MHz - the highest clock BOTH designs close"
    ws["A1"].font = TITLE
    r = 2
    r = para(ws, r, "325 MHz closes on dense (WNS +0.013 ns) and on sparse (+0.001 ns), and "
                    "both are verified bit-exact on hardware at every sparsity. That makes it "
                    "the fairest possible equal-clock comparison. It is NOT a new headline: "
                    "the speed-up is a beat-count ratio and therefore clock-independent, so "
                    "325 confirms 300 rather than replacing it. What it adds is proof that the "
                    "design converts clock into throughput one-for-one.", NOTE, height=44)
    r = para(ws, r, "ESTIMATOR WARNING - the one thing to get right on this sheet. Every "
                    "throughput number here is SUSTAINED, measured over a 60 s soak of ~8,500 "
                    "back-to-back iterations. The 300 MHz headline sheet uses MARGINAL numbers "
                    "from least-squares fits with launch overhead removed. A soak still "
                    "carries that overhead, so it reads ~5-6% lower for identical hardware. "
                    "Comparing a 325 soak against a 300 fit therefore invents a slowdown that "
                    "does not exist. Every cross-clock comparison below is soak against soak.",
             NOTE, height=58)

    r += 1
    ws.cell(row=r, column=1, value="1. POWER AT 325 MHz - all five configurations").font = SEC
    r += 1
    header(ws, r, ["Design", "Board load (W)", "Board idle (W)", "Board dynamic (W)",
                   "VCCINT load (W)", "VCCINT idle (W)", "VCCINT dynamic (W)", "Iterations"],
           widths=[13, 13, 13, 14, 13, 13, 15, 11])
    r += 1
    for lbl in ORDER:
        if lbl not in g325:
            continue
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, avg325(lbl, "board_load_w"), "0.00")
        put(ws, r, 3, avg325(lbl, "board_idle_w"), "0.00")
        put(ws, r, 4, avg325(lbl, "board_delta_w"), "0.00", fill=SUB_FILL)
        put(ws, r, 5, avg325(lbl, "vccint_load_w"), "0.00")
        put(ws, r, 6, avg325(lbl, "vccint_idle_w"), "0.00")
        put(ws, r, 7, avg325(lbl, "vccint_delta_w"), "0.00", fill=SUB_FILL)
        put(ws, r, 8, sum(int(x["iterations"]) for x in g325[lbl]), "#,##0")
        r += 1
    r += 1
    _lo = min(avg325(l, "board_load_w") for l in SPARSE_L)
    _hi = max(avg325(l, "board_load_w") for l in SPARSE_L)
    _si = sum(avg325(l, "board_idle_w") for l in SPARSE_L) / float(len(SPARSE_L))
    _di = avg325("dense", "board_idle_w")
    r = para(ws, r, "POWER IS FLAT ACROSS SPARSITY AGAIN, reproducing the 300 MHz finding: "
                    "board load spans only %.2f-%.2f W (%.1f%%) across a 16x throughput "
                    "range. The engine takes one weight beat per cycle whatever the sparsity, "
                    "so it draws the same watts; throughput rises because a row needs fewer "
                    "beats. Energy per row therefore falls exactly as fast as throughput "
                    "rises. This is the mechanism, not a null result."
                    % (_lo, _hi, 100.0 * (_hi / _lo - 1.0)), NOTE, height=44)
    r = para(ws, r, "STATIC OVERHEAD IS UNCHANGED BY CLOCK, as it must be: sparse idles at "
                    "%.2f W against %.2f W for dense, a gap of %.2f W - the same ~3.2 W "
                    "measured at 300 MHz. Leakage does not care about frequency. Dynamic "
                    "power does, and it duly rose in both designs."
                    % (_si, _di, _si - _di), NOTE, height=30)

    r += 1
    ws.cell(row=r, column=1, value="2. ENERGY PER ROW AT 325 MHz").font = SEC
    r += 1
    header(ws, r, ["Design", "Sustained (Mrow/s)", "Energy/row, board (nJ)", "vs dense",
                   "Energy/row, VCCINT (nJ)", "vs dense", "GMAC/s/W"],
           widths=[13, 16, 17, 11, 17, 11, 12])
    r += 1
    _dm = avg325("dense", "Mrow_s_sustained")
    _dnj = avg325("dense", "energy_per_row_nJ")
    _dnjv = avg325("dense", "vccint_load_w") * 1000.0 / _dm
    for lbl in ORDER:
        if lbl not in g325:
            continue
        m = avg325(lbl, "Mrow_s_sustained")
        nj = avg325(lbl, "energy_per_row_nJ")
        njv = avg325(lbl, "vccint_load_w") * 1000.0 / m
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, m, "0.00")
        put(ws, r, 3, nj, "0.0", font=BOLD)
        put(ws, r, 4, _dnj / nj, XFMT, fill=(HI_FILL if lbl != "dense" else None))
        put(ws, r, 5, njv, "0.0", font=BOLD)
        put(ws, r, 6, _dnjv / njv, XFMT, fill=(OK_FILL if lbl != "dense" else None))
        put(ws, r, 7, avg325(lbl, "GMAC_per_W"), "0.00")
        r += 1
    r += 1
    r = para(ws, r, "Both designs became about 6% more energy-efficient per row simply by "
                    "running faster: the same work spends less time paying the fixed static "
                    "draw. The RATIO between them barely moved (sparse 2:4 is 1.76x better "
                    "than dense here against 1.75x at 300 MHz), which is the expected result "
                    "- energy ratios, like speed-up ratios, are clock-independent.",
             NOTE, height=44)

    r += 1
    ws.cell(row=r, column=1,
            value="3. DOES THROUGHPUT SCALE WITH CLOCK? - soak vs soak, 300 -> 325").font = SEC
    r += 1
    r = para(ws, r, "The decisive table on this sheet, and the reason the soaks were worth "
                    "keeping. Both columns are sustained throughput measured the same way, so "
                    "the ratio is clean. Ideal is 325/300 = 1.0833.", NOTE, height=15)
    header(ws, r, ["Design", "300 MHz (Mrow/s)", "325 MHz (Mrow/s)", "Measured ratio",
                   "Ideal ratio", "Deviation"],
           widths=[13, 16, 16, 13, 12, 11])
    r += 1
    _worst = 0.0
    for lbl in ORDER:
        if lbl not in g325 or lbl not in grouped:
            continue
        a = avg(lbl, "Mrow_s_sustained")
        b = avg325(lbl, "Mrow_s_sustained")
        dev = (b / a) / (325.0 / 300.0) - 1.0
        _worst = max(_worst, abs(dev))
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, a, "0.00")
        put(ws, r, 3, b, "0.00", font=BOLD)
        put(ws, r, 4, b / a, "0.0000")
        put(ws, r, 5, 325.0 / 300.0, "0.0000")
        put(ws, r, 6, dev, "+0.00%;-0.00%", fill=OK_FILL)
        r += 1
    r += 1
    r = para(ws, r, "EVERY configuration is within %.2f%% of ideal linear scaling, dense and "
                    "sparse alike. An 8.33%% clock increase buys 8.33%% more throughput: the "
                    "engine is not limited by anything that fails to scale with the clock. "
                    "This also RETIRES an earlier misreading - the 4-point fits at 325 "
                    "suggested sublinear scaling (1.04-1.07x), which was contention on a "
                    "shared machine, not a property of the design. Section 5 explains how "
                    "that was caught." % (100.0 * _worst), NOTE, height=58)

    r += 1
    ws.cell(row=r, column=1,
            value="4. SPEED-UP vs DENSE AT BOTH CLOCKS - the claim, twice").font = SEC
    r += 1
    r = para(ws, r, "Same estimator at both clocks. If the architecture is what it claims to "
                    "be, these two columns should agree with each other and with theory, and "
                    "the clock should be irrelevant.", NOTE, height=15)
    header(ws, r, ["Sparsity", "Theory", "300 MHz", "325 MHz", "% of theory @300",
                   "% of theory @325"],
           widths=[13, 10, 12, 12, 15, 15])
    r += 1
    _d3 = avg("dense", "Mrow_s_sustained")
    _d32 = avg325("dense", "Mrow_s_sustained")
    for lbl in ORDER[1:]:
        if lbl not in g325 or lbl not in grouped:
            continue
        t = THEORY[lbl]
        a = avg(lbl, "Mrow_s_sustained") / _d3
        b = avg325(lbl, "Mrow_s_sustained") / _d32
        put(ws, r, 1, lbl, font=BOLD)
        put(ws, r, 2, t, XFMT)
        put(ws, r, 3, a, XFMT)
        put(ws, r, 4, b, XFMT, font=BOLD, fill=HI_FILL)
        put(ws, r, 5, a / t, "0.0%")
        put(ws, r, 6, b / t, "0.0%", fill=OK_FILL)
        r += 1
    r += 1
    r = para(ws, r, "The two clock columns agree to about 1%, confirming the speed-up is a "
                    "beat-count ratio and nothing else. These sit a little below the marginal "
                    "figures on the Comparison sheet (1.98 / 3.96 / 7.86 / 15.63x) because a "
                    "soak carries launch overhead that GROWS with output traffic - 2:32 writes "
                    "65,536 output beats per iteration against 8,192 for 2:4 - which penalises "
                    "the sparsest cases most. The fit removes that, which is why the fit "
                    "remains the headline estimator.", NOTE, height=58)

    r += 1
    ws.cell(row=r, column=1,
            value="5. WHY THERE ARE NO MARGINAL FITS AT 325 - a rejected measurement").font = SEC
    r += 1
    r = para(ws, r, "Recorded because the rejection criterion is reusable and because a "
                    "discarded measurement belongs in the record. The 4-point sweeps at 325 "
                    "were run while the host carried a load average near 5 with 10 users "
                    "logged in. The 2:4 sweep returned a fit that is PHYSICALLY IMPOSSIBLE.",
             NOTE, height=30)
    header(ws, r, ["Quantity", "Value", "Verdict"], widths=[38, 15, 44])
    r += 1
    for q, v, fmt, verdict, bad in [
            ("One weight beat per clock at 325 MHz (us)", 1.0 / 325.0, "0.000000",
             "hard floor - the engine cannot beat this", False),
            ("Measured slope, 2:4 sweep at 325 (us)", 0.003064, "0.000000",
             "12.9 ps/beat FASTER than the clock itself", True),
            ("Implied beats per cycle, FIT row", 1.0042, "0.0000",
             "IMPOSSIBLE - the roofline is exactly 1.0000", True),
            ("Fit R^2", 0.99704, "0.00000",
             "worst of the five sweeps; threshold is 0.999", True)]:
        put(ws, r, 1, q, font=BOLD)
        put(ws, r, 2, v, fmt, fill=(HI_FILL if bad else None))
        put(ws, r, 3, verdict)
        r += 1
    r += 1
    r = para(ws, r, "The cause is visible in the per-size launch overhead, which must be "
                    "CONSTANT for the fit to mean anything. Implied overhead across the four "
                    "sizes ran 474.0 / 696.9 / 759.6 / 526.3 us: the 4096-lap point ran slow "
                    "and the 8192-lap point ran fast, and combining a slow point with a fast "
                    "one at the two LARGEST sizes tilts the slope past the clock. Best-of-N "
                    "cannot remove this - keeping the fastest of 15 spans means one lucky "
                    "large sample sets the slope.", NOTE, height=58)
    r = para(ws, r, "Two checks confirmed the design was not at fault. First, a genuinely "
                    "faster-than-target clock would push ALL five configurations below the "
                    "6452.8 us floor at their largest size; none of them are below it, and "
                    "their ordering tracks output traffic exactly as expected. Second, the "
                    "soaks - a completely independent measurement - show textbook linear "
                    "scaling (section 3). The sweep was bad; the silicon was fine.",
             NOTE, height=44)
    r = para(ws, r, "THE REUSABLE RULE, AND ITS SCOPE - this matters, because the naive "
                    "version of the rule is wrong. beats/cycle above 1.000 is always "
                    "physically impossible, but WHICH ROW it appears in decides whether the "
                    "measurement is disqualified:", NOTE, height=30)
    header(ws, r, ["Row type", "Above 1.000 means", "Seen in this corpus"],
           widths=[16, 52, 30])
    r += 1
    for kind, meaning, seen in [
            ("FIT", "DISQUALIFYING - the headline number for that sweep is invalid",
             "1 of 9 sweeps (2:4 at 325, rejected)"),
            ("differential", "routine noise - two sizes had slightly different launch "
             "overhead, and the subtraction exposes the drift",
             "6 of 21 rows, in otherwise sound data"),
            ("raw per-size", "would indicate a broken clock or beat count; never seen",
             "0")]:
        put(ws, r, 1, kind, font=BOLD)
        put(ws, r, 2, meaning)
        put(ws, r, 3, seen, fill=(HI_FILL if kind == "FIT" else None))
        r += 1
    r += 1
    r = para(ws, r, "A scan of all nine throughput CSVs found six differential rows above the "
                    "roofline at 225 and 300 MHz - in data whose FITS are all sound "
                    "(0.958-0.996 beats/cycle, R^2 >= 0.9988). So a single hot differential "
                    "is NOT evidence of a bad sweep, and an earlier note in this project "
                    "claiming three such rows should have been the tell for a bad dataset "
                    "was overstated. What actually disqualifies a sweep is the FIT going "
                    "impossible, which happened exactly once - 2:4 at 325 - and is why that "
                    "one sweep is excluded while every other sweep in this workbook stands.",
             NOTE, height=58)

    # raw 325 rows, unaveraged
    ws3 = wb.create_sheet("power_raw_325")
    ws3.cell(row=1, column=1, value="power_results_325.csv - raw, one row per run, "
                                    "nothing averaged").font = TITLE
    header(ws3, 3, pcols, widths=[13] + [12] * (len(pcols) - 1))
    rr = 4
    for row in p325raw:
        for i, k in enumerate(pcols, start=1):
            if k == "label":
                put(ws3, rr, i, row.get(k, ""), font=BOLD)
            else:
                v = fnum(row, k)
                put(ws3, rr, i, v if v is not None else "",
                    "#,##0" if k == "iterations" else "0.000")
        rr += 1

# ===========================================================================
# 4 -- Methodology
# ===========================================================================
ws = wb.create_sheet("Methodology")
ws.column_dimensions["A"].width = 118
rows = [
    ("GEMV hardware measurement - method and caveats", TITLE),
    ("", None),
    ("PLATFORM", SEC),
    ("Alveo U280 (xcu280-fsvh2892-2L-e), platform xilinx_u280_xdma_201920_3, Vitis/Vivado "
     "2021.1, XRT 2.13. Shared machine - the card is reprogrammed by other users between "
     "sessions.", None),
    ("", None),
    ("SYSTEM UNDER TEST", SEC),
    ("dense : 10 x krnl_mm2s -> krnl_gemv_dense (ap_ctrl_none, free-running) -> 4 x "
     "krnl_s2mm. 14 HBM pseudo-channels. 300 MHz, WNS +0.022 ns.", None),
    ("sparse: 13 x krnl_mm2s -> krnl_gemv_sparse -> 4 x krnl_s2mm. 17 HBM pseudo-channels "
     "(8 weight + 3 index + 2 activation in, 4 out). 300 MHz, WNS 0.000 ns - a genuine pass "
     "(zero failing endpoints) but with NO margin, which should be stated as such.", None),
    ("The data-mover kernels are the SAME BINARIES in both systems - not rebuilt - so the "
     "comparison isolates the compute engine.", None),
    ("", None),
    ("WHY RAW PER-RUN NUMBERS ARE NOT QUOTED", SEC),
    ("A single run is dominated by XRT launch overhead: the first sparse hardware run "
     "measured a 460 us kernel span for TWO weight beats, ~99.99% scheduling. The fitted "
     "intercepts put the fixed overhead at 458-686 us regardless of design.", None),
    ("", None),
    ("METHOD", SEC),
    ("1. Four problem sizes spanning an IDENTICAL beat range in every configuration: "
     "262,144 -> 2,097,152 weight beats.", None),
    ("2. Each size run 15-25 times (sparse) / 5 (dense); the BEST kernel span is kept. Best, "
     "not mean: the machine is shared, so slow samples are contention rather than the "
     "design.", None),
    ("3. Timing from OpenCL profiling events on the kernels (earliest CU start -> latest CU "
     "end), never wall clock - loading an xclbin triggers FPGA reconfiguration, which is "
     "seconds and non-deterministic.", None),
    ("4. A least-squares fit t(beats) = overhead + beats x slope over all four sizes. The "
     "SLOPE is the marginal cost per beat and is what every headline figure derives from; "
     "the INTERCEPT is the fixed launch overhead.", None),
    ("5. Differential rows in the per-design sheets are the same quantity from adjacent size "
     "pairs only. They agree with the fit but are noisier - several report beats/cycle "
     "marginally ABOVE 1.0, which is physically impossible and is best-of estimator "
     "optimism on a single lucky sample, not a real measurement.", None),
    ("", None),
    ("CORRECTNESS", SEC),
    ("Throughput stimulus is synthetic and carries NO golden model - it exists only to "
     "generate beats. That is legitimate because the datapath's timing is data-independent: "
     "the gather is a mux (delay independent of which index it selects), the Xilinx FP "
     "operators are configured with Maximum_Latency, and there is no data-dependent control "
     "flow. Only beat COUNTS affect run length.", None),
    ("Correctness is established separately, on small cases, by the Python golden model - "
     "bit-exact on hardware at all four sparsities AND in a runtime-reconfiguration run "
     "where the sparsity changes three times mid-calculation with no reload. Re-verified on "
     "the 300 MHz bitstream before any number here was recorded. The same golden model "
     "validated behavioural simulation, post-implementation simulation, hardware emulation "
     "and silicon, unchanged throughout.", None),
    ("", None),
    ("A HYPOTHESIS THAT WAS WRONG - recorded so it is not repeated", SEC),
    ("An earlier, noisier sweep (smaller sizes, 5 reps, 30-60% spreads) showed beats/cycle "
     "apparently FALLING with sparsity - 0.969 / 0.831 / 0.879 / 0.787 - and this was "
     "attributed to the output path saturating at 2:32, where the freeze is 1 cycle so a "
     "flush occurs every cycle. That was MEASUREMENT NOISE, not an effect. Re-measured "
     "properly it is flat and near roofline at both clocks. There is no output-path "
     "bottleneck. The tell was present and missed: three differential rows in that data "
     "reported beats/cycle above 1.0, which is impossible.", None),
    ("", None),
    ("POWER MEASUREMENT", SEC),
    ("One calculation lasts 7-8 ms while `xbutil examine` takes ~1 s, so a 1 Hz instrument "
     "cannot sample a single run - and a plain run is ~85% H2D transfer anyway. Both hosts "
     "therefore take a soak argument: re-enqueue every CU back-to-back for 60 s with buffers "
     "ALREADY RESIDENT (no transfers), printing the window as epoch seconds so the sampler "
     "clips to exactly the load. ~54 samples per run survive the 6 s warm-up trim.", None),
    ("The warm-up trim is not cosmetic: on a 5 s trial a sample taken just after the load "
     "began read 28.94 W, BELOW the 29.30 W idle mean, because the card's satellite "
     "controller polls its sensors slowly.", None),
    ("The idle baseline is taken AFTER the host exits, with the same bitstream still "
     "programmed - so leakage, clock trees and the platform are held constant and the "
     "difference isolates the work. The reading taken BEFORE the host runs is NOT a valid "
     "baseline: it reflects whatever design was previously loaded.", None),
    ("Board power = the two 12 V input rails; VCCINT = the FPGA core rail, a SUBSET of it "
     "downstream of the regulators. Verified: 12 V Aux + 12 V PCIe reconstruct the reported "
     "board total exactly. Never sum the three. Only 3 of the U280's 14 rails report current, "
     "so only those three yield power.", None),
    ("Host/CPU power is deliberately NOT measured. The host does identical work in both "
     "cases - same movers, same transfers, same OpenCL calls - so it cannot discriminate "
     "between the designs, and reporting it would add a number that looks like evidence.",
     None),
    ("", None),
    ("KNOWN CAVEATS", SEC),
    ("* The 300 MHz sparse build closes with WNS exactly 0.000 ns - zero failing endpoints, "
     "but zero margin, against dense's +0.022. Correctness was re-verified on it.", None),
    ("* R^2 is 0.9988-0.9994 at 300 MHz, marginally below the 0.999 threshold for 2:16. The "
     "cause is the smallest size in each sweep (25-32% spread from launch-overhead variance "
     "against 4-5% at the largest); more repetitions do not fix it. Independent replicate "
     "runs agreeing to 0.4-1.4% are the better uncertainty estimate - see Clock study.", None),
    ("* V = 1024 elements (32 activation windows) throughout. Vector length was not swept.", None),
    ("* The 225 MHz sparse figures come from a DIFFERENT bitstream (no floorplan). They are "
     "a valid second data point, not an alternative measurement of the same design.", None),
    ("* Power sampling at ~1 Hz measures STEADY STATE only; nothing transient is visible. "
     "Board power also includes fans, which are thermally controlled and drift.", None),
    ("* The dynamic power deltas (4-5 W) are differences of large numbers (~31-35 W) and are "
     "the least precise quantity reported. Standard errors are in the raw CSV and should be "
     "quoted with them. The energy-per-row figures, which divide a LOAD power by a measured "
     "throughput, are far better conditioned.", None),
    ("* Dense was measured twice; dense_300MHz.csv is the run used here. An earlier run "
     "recorded in INTEGRATION_STEPS.md S22 gives a 38.17 Mrow/s DIFFERENTIAL. Use the FIT "
     "(36.99), because every sparse figure is a fit and mixing estimators across the two "
     "sides of a comparison is indefensible. They agree on the physics (~1 beat/cycle).", None),
]
r = 1
for text, font in rows:
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = None if not text else (18 if font in (TITLE, SEC) else 44)
    r += 1

# ===========================================================================
# raw sheets
# ===========================================================================
COLS = [("laps", "Size / pair", 14), ("weight_beats", "Weight beats", 13),
        ("rows", "Rows", 11), ("best_us", "Best us", 11), ("mean_us", "Mean us", 11),
        ("worst_us", "Worst us", 11), ("spread_pct", "Spread %", 10),
        ("reps", "Reps", 8), ("ideal_us", "Ideal us", 11),
        ("bytes_in", "Bytes in", 13), ("bytes_index", "of which index", 13),
        ("bytes_out", "Bytes out", 12), ("Mrow_s", "Mrow/s", 11), ("GB_s", "GB/s", 11),
        ("beats_per_cycle", "beats/cycle", 11)]

for name, path, title in (
        [("dense_300MHz", DENSE[1], "dense_gemv @ 300 MHz")] +
        [("sparse_" + l.replace(":", "to") + "_300MHz", p,
          "two2N sparsity " + l + " @ 300 MHz (floorplanned build)") for l, p in SPARSE300] +
        [("sparse_" + l.replace(":", "to") + "_225MHz", p,
          "two2N sparsity " + l + " @ 225 MHz (pre-floorplan build)") for l, p in SPARSE225]):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title + "  -  source: " + path).font = TITLE
    ws.cell(row=2, column=1,
            value=("Rows 'a->b' are differentials between adjacent sizes (launch overhead "
                   "cancels). Row 'FIT' is the least-squares fit: Best us = FIXED OVERHEAD, "
                   "Mean us = slope in us/beat, Worst us = R^2.")).font = NOTE
    header(ws, 4, [c[1] for c in COLS], widths=[c[2] for c in COLS])
    rr = 5
    for row in load(path):
        is_fit = row["laps"] == "FIT"
        is_diff = "->" in row["laps"]
        fill = HI_FILL if is_fit else (SUB_FILL if is_diff else None)
        for i, (key, _, _) in enumerate(COLS, start=1):
            if key == "laps":
                put(ws, rr, i, row.get(key, ""), font=(BOLD if is_fit else None), fill=fill)
                continue
            num = fnum(row, key)
            if is_fit and key == "mean_us":
                fmt = "0.000000"
            elif is_fit and key == "worst_us":
                fmt = "0.00000"
            elif key in ("weight_beats", "rows", "bytes_in", "bytes_index", "bytes_out"):
                fmt = "#,##0"
            elif key == "reps":
                fmt = "0"
            else:
                fmt = "0.000"
            put(ws, rr, i, num if num is not None else "", fmt,
                font=(BOLD if is_fit else None), fill=fill)
        rr += 1

wb.save(OUT)
print("wrote", OUT)
print("sheets:", ", ".join(wb.sheetnames))
