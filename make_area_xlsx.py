"""Build GEMV_Area_Comparison.xlsx from the two OOC utilization reports.

WHAT THIS COMPARES, AND WHAT IT DOES NOT. Both numbers come from MODULE-LEVEL
OUT-OF-CONTEXT implementation of the compute engine alone, on the same part
(xcu280-fsvh2892-2L-e) under the same 2.222 ns / 450 MHz constraint, with no
synthesis strategy applied to either. The HLS data movers, the AXI interconnect
and the platform shell are NOT included -- this isolates what the sparsity
itself costs. For what the whole accelerator costs on the device, use the
system-level reports (impl_1_kernel_util_routed.rpt) instead, and never mix the
two in one table: they differ by 17 mover kernels and the shell.

WRAPPER TO WRAPPER. Both sides are the AXI-Stream wrapped variants
(two2N_axis, dense_gemv_axis) so the comparison is like for like. The wrapper
was separately shown to be free: two2N_axis came out 18 LUTs and 40 FFs SMALLER
than bare two2N with F7/DSP/BRAM exactly equal, which is OOC-boundary tool
variation rather than a real saving.

SIGN CONVENTION -- THIS MATTERS. Dense is the baseline, so every delta is what
SPARSE COSTS relative to it, and the percentages use dense as denominator.
Sparse is 41.1% LARGER in LUTs. An earlier note recorded this as "-29%", which
is the same fact with sparse as the denominator and reads, wrongly, as though
sparsity saved area. State it as a cost.

Run:  python make_area_xlsx.py
"""

import io
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "results")


def _data(name):
    # results/ first, then the repo root -- so a CSV freshly copied down from
    # the server is still found before it has been filed away.
    p = os.path.join(DATA, name)
    return p if os.path.exists(p) else os.path.join(HERE, name)


def _out(name):
    if not os.path.isdir(DATA):
        os.makedirs(DATA)
    return os.path.join(DATA, name)


OUT = _out("GEMV_Area_Comparison.xlsx")
SPARSE = "two2N_axis_utilization_placed.rpt"
DENSE = "dense_gemv_axis_utilization_placed.rpt"

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(color="FFFFFF", bold=True, size=11)
SEC_FILL = PatternFill("solid", fgColor="D9E2F3")
HI = PatternFill("solid", fgColor="FFF2CC")
OK = PatternFill("solid", fgColor="E2EFDA")
TITLE = Font(bold=True, size=14, color="1F3864")
SEC = Font(bold=True, size=11, color="1F3864")
NOTE = Font(italic=True, size=9, color="555555")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="C8C8C8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# The rows that carry the argument, in the order they should be read.
HEADLINE = [
    ("CLB LUTs", "the headline area cost of per-row sparsity"),
    ("LUT as Logic", "all of the growth is here"),
    ("LUT as Memory", "identical -- the shift-register delay lines are unchanged"),
    ("CLB Registers", "the activation window, replicated per core to fix fanout"),
    ("CARRY8", "identical -- no extra arithmetic carry chains"),
    ("F7 Muxes", "THE GATHER: 64 blocks x 2 indices x 16 bits x 4 MUXF7"),
    ("Block RAM Tile", "11 input pseudo-channels vs 8, plus the replay buffer"),
    ("DSPs", "IDENTICAL -- sparsity adds selection, never arithmetic"),
]
# Where the LUTs went, and proof the datapath is unchanged.
DETAIL = [
    ("LUT6", "the gather is 6-input multiplexing: this IS the whole delta"),
    ("LUT5", ""),
    ("LUT4", ""),
    ("LUT3", ""),
    ("LUT2", ""),
    ("LUT1", ""),
    ("FDRE", ""),
    ("SRL16E", "identical"),
    ("RAMB36E2", ""),
    ("DSP48E2", "identical"),
    ("Multiplier", "IP instance count -- identical"),
    ("Adder", "IP instance count -- identical"),
    ("Accumulator", "IP instance count -- identical"),
]


def parse(path):
    """Site type / primitive -> count, from a Vivado utilization report."""
    s = io.open(_data(path), encoding="utf-8", errors="replace").read()
    # The per-SLR tables repeat every site type; stop before them. Use rfind so
    # the table of contents near the top does not truncate the whole report.
    cut = s.rfind("12. SLR Connectivity")
    if cut > 0:
        s = s[:cut]
    out = {}
    for line in s.splitlines():
        m = re.match(r"\|\s*([A-Za-z0-9 /\-\.]+?)\s*\|\s*([\d.]+)\s*\|", line)
        if m:
            name = m.group(1).strip()
            if name and name not in ("Site Type", "Ref Name") and name not in out:
                out[name] = float(m.group(2))
    meta = {}
    for key in ("Design", "Date", "Device"):
        mm = re.search(r"\|\s*" + key + r"\s*:\s*(.+)", s)
        meta[key] = mm.group(1).strip() if mm else "?"
    return out, meta


def put(ws, r, c, v, fmt=None, font=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.border = BOX
    return cell


def para(ws, r, text, height=30):
    c = ws.cell(row=r, column=1, value=text)
    c.font = NOTE
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = height
    return r + 1


def header(ws, r, cols, widths):
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=name)
        c.fill, c.font, c.border = H_FILL, H_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 30
    return r + 1


def block(ws, r, spec, S, D):
    for name, note in spec:
        if name not in S or name not in D:
            continue
        a, b = S[name], D[name]
        d = a - b
        put(ws, r, 1, name, font=BOLD)
        put(ws, r, 2, int(a), "#,##0")
        put(ws, r, 3, int(b), "#,##0")
        put(ws, r, 4, int(d), "+#,##0;-#,##0;0",
            fill=(HI if d > 0 else (OK if d == 0 else None)))
        if b:
            put(ws, r, 5, d / b, "+0.0%;-0.0%",
                fill=(HI if d > 0 else (OK if d == 0 else None)))
        else:
            put(ws, r, 5, "new" if d else "", fill=(HI if d else None))
        put(ws, r, 6, note)
        r += 1
    return r


def main():
    S, ms = parse(SPARSE)
    D, md = parse(DENSE)

    wb = Workbook()
    ws = wb.active
    ws.title = "Area"
    ws["A1"] = "Resource comparison - sparse 2:M engine vs dense baseline"
    ws["A1"].font = TITLE
    r = 2
    r = para(ws, r,
             "MODULE-LEVEL OUT-OF-CONTEXT implementation of the compute engine ALONE. Same "
             "part (%s), same 2.222 ns / 450 MHz constraint, OOC mode, no synthesis strategy "
             "on either side. The 17 HLS data movers, the AXI interconnect and the platform "
             "shell are NOT included -- this isolates what the sparsity itself costs. For what "
             "the whole accelerator costs on the device, use the system-level reports instead, "
             "and never mix the two in one table." % md.get("Device", "xcu280"), height=44)
    r = para(ws, r,
             "SIGN CONVENTION: dense is the baseline. Every delta is what SPARSE COSTS, and "
             "percentages use dense as the denominator. Sparse is 41.1%% LARGER in LUTs. "
             "Quoting this as \"-29%%\" (sparse as denominator) reads as though sparsity saved "
             "area, which it does not.", height=30)
    r = para(ws, r,
             "Both sides are the AXI-Stream wrapped variants (%s vs %s), so the comparison is "
             "like for like. The wrapper was separately shown to be free: two2N_axis came out "
             "18 LUTs and 40 FFs SMALLER than bare two2N with F7/DSP/BRAM exactly equal -- "
             "OOC-boundary tool variation, not a real saving."
             % (ms.get("Design"), md.get("Design")), height=30)
    r = para(ws, r,
             "PROVENANCE. Sparse report: %s. Dense report: %s. The dense run predates the "
             "end-of-calculation teardown fix (written 2026-08-23) by one RTL change; the "
             "equivalent fixes on sparse cost +65 LUTs, i.e. 0.1%%. Re-run dense OOC for a "
             "same-vintage pair if the asterisk matters."
             % (ms.get("Date"), md.get("Date")), height=30)

    r += 1
    ws.cell(row=r, column=1, value="1. HEADLINE RESOURCES").font = SEC
    r += 1
    r = header(ws, r, ["Resource", "Sparse", "Dense", "Sparse cost", "vs dense",
                       "What it means"], [22, 12, 12, 13, 11, 62])
    r = block(ws, r, HEADLINE, S, D)

    r += 1
    ws.cell(row=r, column=1, value="2. WHERE THE LUTs WENT, AND WHAT DID NOT CHANGE").font = SEC
    r += 1
    r = header(ws, r, ["Primitive", "Sparse", "Dense", "Sparse cost", "vs dense",
                       "Note"], [22, 12, 12, 13, 11, 62])
    r = block(ws, r, DETAIL, S, D)

    r += 1
    r = para(ws, r,
             "THE ENTIRE LUT DELTA IS LUT6: +18,514 against a net +18,176. LUT1-5 move by "
             "roughly a hundred in total. The gather is 6-input multiplexing and nothing else "
             "changed.", height=30)
    r = para(ws, r,
             "AND THE COMPUTE DATAPATH IS INSTANCE-FOR-INSTANCE IDENTICAL: 128 Multipliers, "
             "64 Adders, 64 Accumulators, 512 DSP48E2, 6,784 SRL16E in BOTH designs. This is "
             "stronger than \"the DSP count matches\" -- the arithmetic is literally the same "
             "hardware. Sparsity buys its speed-up by SELECTING different operands, never by "
             "adding arithmetic.", height=44)

    r += 1
    ws.cell(row=r, column=1, value="3. WHAT THE AREA BUYS (300 MHz, measured)").font = SEC
    r += 1
    r = header(ws, r, ["Configuration", "GFLOPS eff", "kLUT", "GFLOPS / kLUT",
                       "vs dense", ""], [22, 12, 12, 13, 11, 62])
    klut_s, klut_d = S["CLB LUTs"] / 1000.0, D["CLB LUTs"] / 1000.0
    base = 69.31 / klut_d
    for lab, gf, kl in (("dense", 69.31, klut_d), ("sparse 2:4", 137.19, klut_s),
                        ("sparse 2:8", 273.60, klut_s), ("sparse 2:16", 542.92, klut_s),
                        ("sparse 2:32", 1073.04, klut_s), ("sparse MIXED", 288.45, klut_s)):
        put(ws, r, 1, lab, font=BOLD)
        put(ws, r, 2, gf, "0.0")
        put(ws, r, 3, kl, "0.0")
        put(ws, r, 4, gf / kl, "0.00", font=BOLD)
        put(ws, r, 5, (gf / kl) / base, '0.00"x"',
            fill=(HI if lab != "dense" else None))
        put(ws, r, 6, "")
        r += 1
    r += 1
    r = para(ws, r,
             "41%% more LUTs buys up to 11x more work per LUT. That is the trade, and it is "
             "the sentence for the slide. Effective GFLOPS are the mean-of-3 hardware "
             "measurements at 300 MHz; MIXED is one matrix carrying all four sparsities.",
             height=30)

    wb.save(OUT)
    print("wrote", OUT)
    print("  sparse: %s  (%s)" % (ms.get("Design"), ms.get("Date")))
    print("  dense : %s  (%s)" % (md.get("Design"), md.get("Date")))


if __name__ == "__main__":
    main()
