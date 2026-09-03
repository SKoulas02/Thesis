"""Build GEMV_Chart_Data.xlsx -- the result CSVs as clean, chartable sheets.

This is DELIBERATELY not GEMV_Hardware_Results.xlsx. That workbook is the
annotated record: narrative, caveats, method notes, nineteen sheets. This one is
raw chartable data and nothing else, so that selecting columns and pressing
"Insert Chart" does the right thing.

    GEMV_avg3_results.csv       -> "avg3 comparison" PRIMARY: dense + 4 sparsities
                                                     + mixed, both clocks, ONE
                                                     estimator throughout
    GEMV_shapes_300MHz.csv      -> "Shapes"          11 matrix shapes x 5 configs,
                                                     with GEOMEAN rows
    GEMV_Master_Results.csv     -> "Master"          the fuller record (sweep fits
                                                     + power soaks)
    mixed_sparsity_300MHz.csv   -> "Mixed sparsity"  the runtime-reconfig run
    GEMV_Floorplan_SLR.csv      -> "Floorplan SLR"   per-die resources

What matters for charting, and what this script guarantees:

  * NUMBERS ARE NUMERIC CELLS. A CSV opened directly often lands numbers as
    text, and Excel silently plots text as zero. Every value that parses as a
    number is written as a number here.
  * BLANKS STAY BLANK, never zero. A power_soak row has no bandwidth; a
    sweep_fit row has no watts. Excel skips blanks in a chart but plots zeros,
    which would put a false trough in every series.
  * Header row is frozen and carries an AutoFilter, because the single most
    important operation on the Master sheet is filtering by `estimator`.

    THE ESTIMATOR COLUMN IS LOAD-BEARING. sweep_fit numbers are marginal
    (launch overhead removed by a least-squares fit), power_soak numbers are
    sustained (overhead included), mixed_run_avg3 is the mean of three single
    runs. They differ by 5-6% for IDENTICAL hardware. Charting one against
    another invents a slowdown that does not exist. Filter first.

Run:  python make_chart_xlsx.py
"""

import csv
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "results")


def _data(name):
    # results/ first, then the repo root -- so a CSV freshly copied down from
    # the server is still found before it has been filed away.
    p = os.path.join(DATA, name)
    return p if os.path.exists(p) else os.path.join(HERE, name)


def _out(name):
    if not os.path.isdir(DATA):
        os.makedirs(DATA)
    return os.path.join(DATA, name)


OUT = _out("GEMV_Chart_Data.xlsx")

SOURCES = [
    # FIRST because it is the primary comparison: dense and all four sparsities
    # plus the mixed matrix, at both clocks, ALL measured in one session with a
    # single estimator (mean of 3+ runs). Nothing here needs an estimator
    # caveat, so it is the sheet to chart from.
    ("GEMV_avg3_results.csv", "avg3 comparison"),
    # the matrix-shape sweep: 11 GEMV shapes x 5 configurations, each shape
    # batched so the launch overhead is a subtractable fraction rather than
    # the whole measurement. GEOMEAN rows carry the summary bar.
    ("GEMV_shapes_300MHz.csv", "Shapes"),
    # energy per matrix, split static/dynamic: measured latency x measured
    # power. Use the per-ELEMENT columns for charting -- absolute energy
    # spans 864x across these shapes and cannot be stacked on a log axis.
    ("GEMV_Energy_300MHz.csv", "Energy"),
    ("GEMV_Master_Results.csv", "Master"),
    ("mixed_sparsity_300MHz.csv", "Mixed sparsity"),
    ("GEMV_Floorplan_SLR.csv", "Floorplan SLR"),
]

# Columns that hold labels even when they look numeric, so they are never
# coerced. "sparsity" is the clearest case: 2:4 is a name, not a ratio, and
# some spreadsheets will read it as a time value if given the chance.
TEXT_COLS = {
    "design", "sparsity", "sparsity_code", "estimator", "row_type",
    "site_type", "holds",
}

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D0D0D0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def as_number(text):
    """Return int/float for a numeric string, else None."""
    t = (text or "").strip()
    if not t:
        return None
    try:
        f = float(t)
    except ValueError:
        return None
    # keep whole numbers as ints so Excel shows 4096 rather than 4096.0
    if f.is_integer() and abs(f) < 2 ** 53:
        return int(f)
    return f


def write_sheet(wb, path, title):
    with io.open(_data(path), newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return None
    cols = list(rows[0].keys())
    ws = wb.create_sheet(title)

    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font, c.border = H_FILL, H_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    widest = [len(c) for c in cols]
    for r, row in enumerate(rows, start=2):
        for i, key in enumerate(cols, start=1):
            raw = (row.get(key) or "").strip()
            if not raw:
                continue                      # leave truly blank -- never 0
            num = None if key in TEXT_COLS else as_number(raw)
            cell = ws.cell(row=r, column=i, value=raw if num is None else num)
            cell.border = BOX
            if num is not None and isinstance(num, int) and abs(num) >= 1000:
                cell.number_format = "#,##0"
            widest[i - 1] = max(widest[i - 1], len(raw))

    for i, w in enumerate(widest, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(30, max(10, w + 2))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(cols)), len(rows) + 1)
    return len(rows), len(cols)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    made = []
    for path, title in SOURCES:
        if not os.path.exists(_data(path)):
            print("MISSING, skipped:", path)
            continue
        got = write_sheet(wb, path, title)
        if got:
            made.append((title, path, got[0], got[1]))
    wb.save(OUT)

    print("wrote", OUT)
    for title, path, nrow, ncol in made:
        print("  %-16s <- %-28s %3d rows x %2d cols" % (title, path, nrow, ncol))


if __name__ == "__main__":
    main()
